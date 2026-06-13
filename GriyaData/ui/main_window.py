# main_window.py
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView,
    QPushButton, QLabel, QFrame, QComboBox, QDialog,
    QMessageBox, QSizePolicy, QLineEdit, QFormLayout,
    QDialogButtonBox, QSpacerItem, QSizePolicy as QSP, QToolButton,
    QMenu
)
from PySide6.QtCore import Qt, Slot, QThread, Signal, QTimer, QPoint
from PySide6.QtGui import QAction, QColor, QFont

import csv
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from PySide6.QtWidgets import QFileDialog

from core.api_handler import APIHandler, OrderRecord, ProductRecord
from utils import Formatter
from ui.chart_widget import ChartWidget
from ui.dialog_order import DialogOrder
from ui.dialog_import import DialogImport
from ui.tab_prediksi import TabPrediksi

ANGGOTA = [
    ("I Nyoman Swardi Jaya Putra",  "F1D02310057"),
    ("Ahmad Ridho Fakhrezi",    "F1D021025"),
    ("Irfan Jayadi",   "F1D02310011"),
]
_ANGGOTA_STR = "  |  ".join(f"{n} ({nim})" for n, nim in ANGGOTA)

APP_TITLE = "GriyaData — Manajemen Penjualan"

class FilterWorker(QThread):
    finished = Signal(object)
    error = Signal(str)

    def __init__(self, items, mode, params):
        super().__init__()
        self.items = items
        self.mode = mode
        self.params = params

    def run(self):
        try:
            if self.mode == "orders":
                res = self._filter_orders(self.items, self.params)
            else:
                res = self._filter_products(self.items, self.params)
            self.finished.emit(res)
        except Exception as e:
            self.error.emit(str(e))

    def _filter_orders(self, orders, p):
        out = list(orders)
        col_key = p.get("col_key")
        col_type = p.get("col_type", "text")
        op = p.get("op", "All")
        val = p.get("value")
        kw = (p.get("keyword") or "").strip().lower()
        exact = bool(p.get("exact", False))

        if col_type == "number" and val is not None and op != "All":
            filtered = []
            for o in out:
                try:
                    # try numeric attribute first; fallback to 0
                    num = float(getattr(o, col_key, 0) or 0)
                except Exception:
                    continue
                if op == "<" and num < val: filtered.append(o)
                if op == "=" and num == val: filtered.append(o)
                if op == ">" and num > val: filtered.append(o)
            out = filtered
        # text filter (robust: cast attribute to string)
        elif col_type == "text" and kw:
            if exact:
                out = [o for o in out if str(getattr(o, col_key, "") or "").strip().lower() == kw]
            else:
                out = [o for o in out if kw in str(getattr(o, col_key, "") or "").lower()]
        return out

    def _filter_products(self, prods, p):
        out = list(prods)
        col_key = p.get("col_key")
        col_type = p.get("col_type", "text")
        op = p.get("op", "All")
        val = p.get("value")
        kw = (p.get("keyword") or "").strip().lower()
        category = p.get("category", None)

        if col_key == "category" and category and category != "All":
            out = [x for x in out if (x.category or "") == category]
        elif col_type == "number" and val is not None and op != "All":
            filtered = []
            for x in out:
                try:
                    num = float(getattr(x, col_key, 0) or 0)
                except Exception:
                    continue
                if op == "<" and num < val: filtered.append(x)
                if op == "=" and num == val: filtered.append(x)
                if op == ">" and num > val: filtered.append(x)
            out = filtered
        elif col_type == "text" and kw:
            out = [x for x in out if kw in str(getattr(x, col_key, "") or "").lower()]
        return out

class DataLoader(QThread):
    finished = Signal(list)
    error    = Signal(str)
    def __init__(self, api):
        super().__init__()
        self.api = api
    def run(self):
        try:
            self.finished.emit(self.api.get_all_orders("All", "All"))
        except Exception as e:
            self.error.emit(str(e))

class DialogProduk(QDialog):
    def __init__(self, parent=None, product: ProductRecord = None):
        super().__init__(parent)
        self.setWindowTitle("Tambah Produk" if product is None else "Edit Produk")
        self.setMinimumWidth(360)
        self.setObjectName("dialogSales")
        self._build(product)

    def _build(self, p):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 20, 24, 16); lay.setSpacing(14)

        title = QLabel("Tambah Produk Baru" if p is None else f"Edit Produk ID {p.id}")
        title.setObjectName("dialogTitle")
        title.setStyleSheet("font-size:14px; font-weight:700;")
        lay.addWidget(title)

        form = QFormLayout(); form.setSpacing(10)
        self.inp_name = QLineEdit(p.product_name if p else "")
        self.inp_name.setObjectName("inputField")
        self.inp_name.setPlaceholderText("e.g. Meja Belajar")
        self.inp_cat  = QLineEdit(p.category if p else "")
        self.inp_cat.setObjectName("inputField")
        self.inp_cat.setPlaceholderText("e.g. Meja")
        self.inp_price = QLineEdit(str(p.price) if p else "")
        self.inp_price.setObjectName("inputField")
        self.inp_price.setPlaceholderText("e.g. 150000")

        form.addRow("Nama Produk *", self.inp_name)
        form.addRow("Category *",    self.inp_cat)
        form.addRow("Price *",       self.inp_price)
        lay.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText("Simpan")
        btns.button(QDialogButtonBox.Ok).setObjectName("btnPrimary")
        btns.button(QDialogButtonBox.Cancel).setObjectName("btnDanger")
        btns.accepted.connect(self._validate)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

    def _validate(self):
        if not self.inp_name.text().strip():
            QMessageBox.warning(self, "Input Kurang", "Nama produk tidak boleh kosong."); return
        if not self.inp_cat.text().strip():
            QMessageBox.warning(self, "Input Kurang", "Category tidak boleh kosong."); return
        try:
            float(self.inp_price.text().replace(",",""))
        except ValueError:
            QMessageBox.warning(self, "Input Salah", "Price harus berupa angka."); return
        self.accept()

    def get_data(self) -> dict:
        return {
            "product_name": self.inp_name.text().strip(),
            "category":     self.inp_cat.text().strip(),
            "price":        float(self.inp_price.text().replace(",","").strip() or 0),
        }
class MainWindow(QMainWindow):
    def __init__(self, username="Admin", user_id=None):
        super().__init__()
        self.username = username
        self.user_id  = user_id
        self.setWindowTitle(APP_TITLE)
        self.setMinimumSize(1280, 760)
        self.showMaximized()

        self.api = APIHandler()
        self.api.current_user_id = user_id
        self._orders: list[OrderRecord] = []
        self._loader = None

        self._prod_timer = QTimer(self); self._prod_timer.setSingleShot(True); self._prod_timer.setInterval(300)
        self._order_timer = QTimer(self); self._order_timer.setSingleShot(True); self._order_timer.setInterval(300)

        self._running_workers = []

        self._unique_cache: dict[str, list[str]] = {}

        self._order_enum_columns = {
            "product_name",
            "status",
            "category",
            "payment_method",
            "courier",
            "sales_channel",
            "customer_city",
            "customer_gender",
        }
        self._product_enum_columns = {
            "category",
        }

        self._build_menu()
        self._build_ui()
        self._build_statusbar()
        self.refresh_all()

    def _mark_invalid(self, widget: QLineEdit, msg: str | None = None):
        widget.setStyleSheet("border: 1.5px solid #ef4444;")  # red
        if msg:
            self.lbl_status.setText(msg)

    def _mark_valid(self, widget: QLineEdit):
        widget.setStyleSheet("")  # reset
        if "angka" in self.lbl_status.text().lower():
            self.lbl_status.setText("")

    def _validate_and_parse_number(self, widget: QLineEdit):
        txt = widget.text().strip()
        if not txt:
            self._mark_valid(widget)
            return None
        try:
            val = float(txt.replace(",", ""))
            self._mark_valid(widget)
            return val
        except ValueError:
            self._mark_invalid(widget, "Input angka tidak valid; abaikan filter numerik.")
            return None

    def _build_menu(self):
        mb = self.menuBar(); mb.setObjectName("menuBar")

        # ── File
        mf = mb.addMenu("&File")
        r_data = QAction("Refresh Data", self)
        r_data.setShortcut("F5")
        r_data.triggered.connect(self.refresh_all)
        mf.addAction(r_data)

        # ── Pesanan
        md = mb.addMenu("&Pesanan")
        for label, shortcut, slot in [
            ("Tambah Data Pesanan", "Ctrl+N", self._tambah_order),
            ("Edit Pesanan",        "Ctrl+E", self._edit_order),
            ("Hapus Pesanan",       "Delete", self._hapus_order),
            ("Import CSV/Excel",    "Ctrl+I", self._import_file),
            ("Export CSV",          "Ctrl+Shift+C", self._export_csv),
            ("Export Excel",        "Ctrl+Shift+E", self._export_excel),
        ]:
            a = QAction(label, self); a.setShortcut(shortcut)
            a.triggered.connect(slot); md.addAction(a)
        md.addSeparator()

        # ── Help 
        mh = mb.addMenu("&Help")
        a_about = QAction("Tentang GriyaData", self)
        a_about.triggered.connect(self._show_about)
        a_filter = QAction("Petunjuk Filter & Pencarian", self)
        a_filter.triggered.connect(lambda: self._show_help("orders"))
        mh.addAction(a_about)
        mh.addAction(a_filter)
        
        # ── Exit
        me = mb.addMenu("&Exit")
        for label, shortcut, slot in [
            ("Logout",         "Ctrl+L", self._logout),
            ("Keluar",         "Ctrl+Q", self.close),
        ]:
            a = QAction(label, self); a.setShortcut(shortcut)
            a.triggered.connect(slot); me.addAction(a)

    def _build_ui(self):
        central = QWidget(); central.setObjectName("mainContainer")
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(0,0,0,0); root.setSpacing(0)
        root.addWidget(self._make_banner())
        self.tabs = QTabWidget(); self.tabs.setObjectName("mainTabs")
        root.addWidget(self.tabs, 1)
        self._build_tab_dashboard()
        self._build_tab_orders()
        self._build_tab_products()
        self._build_tab_prediksi()

    def _make_banner(self):
        b = QFrame(); b.setObjectName("banner")
        lay = QHBoxLayout(b); lay.setContentsMargins(20,12,20,12)
        lbl = QLabel("GriyaData — Dashboard Manajemen Furniture")
        lbl.setObjectName("bannerApp")
        lbl.setStyleSheet("font-size:19px;font-weight:700;color:#1a1a1a;")
        lay.addWidget(lbl); lay.addStretch()
        fr = QFrame(); fr.setObjectName("frameIdentitas")
        fr.setMinimumWidth(280)
        il = QHBoxLayout(fr); il.setContentsMargins(14,6,14,6); il.setSpacing(16)
        il.addWidget(QLabel(f"User: {self.username}"))
        il.addWidget(QLabel("API: griyadataapi"))
        lay.addWidget(fr)
        return b

    def _build_tab_dashboard(self):
        tab = QWidget(); tab.setObjectName("tabDashboard")
        root = QVBoxLayout(tab); root.setContentsMargins(20,16,20,16); root.setSpacing(16)
        cl = QHBoxLayout(); cl.setSpacing(12)
        self._card_tx    = self._stat_card("Total Pesanan",      "—", "#3b82f6")
        self._card_rev   = self._stat_card("Total Sales",        "—", "#10b981")
        self._card_avg   = self._stat_card("Rata-rata / Order",  "—", "#f59e0b")
        self._card_units = self._stat_card("Total Unit Terjual", "—", "#8b5cf6")
        for c in (self._card_tx, self._card_rev, self._card_avg, self._card_units):
            cl.addWidget(c)
        root.addLayout(cl)
        self._dash_chart = ChartWidget()
        try:
            self._dash_chart.combo_chart.setCurrentIndex(2)
        except Exception:
            pass
        root.addWidget(self._dash_chart, 1)
        self.tabs.addTab(tab, "Dashboard")

    def _stat_card(self, title, value, color):
        card = QFrame(); card.setObjectName("statCard")
        card.setStyleSheet("""
            #statCard{background:#ffffff;border:1.5px solid #e5e7eb;border-radius:10px;}
            #statCard:hover{border-color:#bfdbfe;}""")
        lay = QVBoxLayout(card); lay.setContentsMargins(18,14,18,14); lay.setSpacing(4)
        lv = QLabel(value); lv.setStyleSheet(f"font-size:24px;font-weight:700;color:{color};")
        lv.setObjectName("cardVal")
        lt = QLabel(title); lt.setStyleSheet("font-size:11px;color:#6b7280;font-weight:600;")
        lay.addWidget(lv); lay.addWidget(lt)
        return card

    def _set_card(self, card, val):
        lbl = card.findChild(QLabel, "cardVal")
        if lbl: lbl.setText(val)

    def _build_tab_orders(self):
        tab = QWidget(); tab.setObjectName("tabTable")
        root = QVBoxLayout(tab); root.setContentsMargins(16,14,16,14); root.setSpacing(8)

        top = QHBoxLayout(); top.setSpacing(8)

        left = QHBoxLayout(); left.setSpacing(6)
        self.order_col_cb = QComboBox()
        self._order_filter_cols = [
            ("ID DB", "id", "number"),
            ("Order ID", "order_id", "text"),
            ("Pelanggan", "customer_name", "text"),
            ("Produk", "product_name", "text"),
            ("Category", "category", "text"),
            ("Price", "price", "number"),
            ("Qty", "quantity", "number"),
            ("Discount", "discount", "number"),
            ("Total", "total", "number"),
            ("Shipping Fee", "shipping_fee", "number"),
            ("Total Sales", "total_sales", "number"),
            ("Status", "status", "text"),
            ("Alamat", "shipping_address", "text"),
            ("Gender", "customer_gender", "text"),
            ("Kota", "customer_city", "text"),
            ("Payment", "payment_method", "text"),
            ("Courier", "courier", "text"),
            ("Est. Hari", "estimated_delivery_days", "number"),
            ("Channel", "sales_channel", "text"),
            ("Rating", "customer_rating", "number"),
            ("Sales Date", "sales_date", "text"),
        ]
        for label, key, typ in self._order_filter_cols:
            self.order_col_cb.addItem(label, (key, typ))
        self.order_col_cb.currentIndexChanged.connect(self._on_order_col_changed)
        left.addWidget(QLabel("Kolom:")); left.addWidget(self.order_col_cb)

        self.order_value_cb = QComboBox()
        self.order_value_cb.setVisible(False)
        self.order_value_cb.setObjectName("inputField")
        self.order_value_cb.currentTextChanged.connect(lambda _: self._order_timer.start())
        left.addWidget(self.order_value_cb)

        self.order_num_op = QComboBox(); self.order_num_op.addItems(["All","<","=" ,">"])
        self.order_num_op.setMinimumWidth(80); self.order_num_op.currentTextChanged.connect(lambda _: self._order_timer.start())
        self.order_num_input = QLineEdit(); self.order_num_input.setPlaceholderText("Angka")
        self.order_num_input.setObjectName("inputField")

        self.order_num_input.textChanged.connect(lambda _: (self._validate_and_parse_number(self.order_num_input), self._order_timer.start()))
        left.addWidget(self.order_num_op); left.addWidget(self.order_num_input)

        top.addLayout(left)

        top.addItem(QSpacerItem(20, 10, QSP.Expanding, QSP.Minimum))

        right = QHBoxLayout(); right.setSpacing(8)
        self.order_text_search = QLineEdit(); self.order_text_search.setPlaceholderText("Kata kunci (untuk kolom teks)")
        self.order_text_search.setObjectName("inputField")
        self.order_text_search.textChanged.connect(lambda _: self._order_timer.start())
        self._order_timer.timeout.connect(self._start_order_worker)
        right.addWidget(self.order_text_search)

        btn_help = QToolButton(); btn_help.setText("?"); btn_help.setToolTip("Petunjuk penggunaan filter dan kata kunci")
        btn_help.clicked.connect(lambda: self._show_help("orders"))
        right.addWidget(btn_help)

        btn_add = QPushButton("Tambah Data Pesanan"); btn_add.setObjectName("btnPrimary"); btn_add.clicked.connect(self._tambah_order)
        right.addWidget(btn_add)

        self.btn_edit = QPushButton("Edit"); self.btn_edit.setEnabled(False); self.btn_edit.clicked.connect(self._edit_order)
        self.btn_edit.setObjectName("btnSecondary"); right.addWidget(self.btn_edit)
        self.btn_del = QPushButton("Hapus"); self.btn_del.setEnabled(False); self.btn_del.clicked.connect(self._hapus_order)
        self.btn_del.setObjectName("btnDanger"); right.addWidget(self.btn_del)

        btn_export = QPushButton("Export")
        btn_export.setObjectName("btnSecondary")
        exp_menu = QMenu(self)
        exp_menu.addAction("Export CSV", self._export_csv)
        exp_menu.addAction("Export Excel", self._export_excel)
        btn_export.setMenu(exp_menu)
        right.addWidget(btn_export)

        top.addLayout(right)

        root.addLayout(top)

        # Table
        self.table = QTableWidget(); self.table.setObjectName("dataTable")
        self._ORDER_COLS = [
            "ID DB", "Order ID", "Pelanggan", "Produk", "Category", "Price",
            "Qty", "Discount", "Total", "Shipping Fee", "Total Sales",
            "Status", "Alamat", "Gender", "Kota", "Payment",
            "Courier", "Est. Hari", "Channel", "Rating", "Sales Date",
        ]
        self.table.setColumnCount(len(self._ORDER_COLS))
        self.table.setHorizontalHeaderLabels(self._ORDER_COLS)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setMinimumSectionSize(60)
        hdr.sectionClicked.connect(self._on_header_clicked)  # header click -> menu for enum columns
        self.table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(False)
        self.table.setStyleSheet("""
            QTableWidget::item:selected {
                background-color: #3b82f6; 
                color: white;
            }
        """)
        self.table.itemSelectionChanged.connect(self._on_sel_changed)
        self.table.doubleClicked.connect(self._edit_order)
        root.addWidget(self.table)
        self.tabs.addTab(tab, "Data Pesanan")
        
        self._on_order_col_changed(0)

    def _on_header_clicked(self, logicalIndex: int):
        if logicalIndex < 0 or logicalIndex >= len(self._order_filter_cols):
            return

        try:
            self.order_col_cb.blockSignals(True)
            self.order_col_cb.setCurrentIndex(logicalIndex)
        finally:
            self.order_col_cb.blockSignals(False)

        self._on_order_col_changed(logicalIndex)

        label, col_key, col_type = self._order_filter_cols[logicalIndex]

        if col_key not in self._order_enum_columns:
            return

        vals = self._unique_cache.get(col_key)
        if vals is None:
            vals = sorted({(str(getattr(o, col_key, "") or "")).strip() for o in self._orders if getattr(o, col_key, None) not in (None, "")})
            self._unique_cache[col_key] = vals

        menu = QMenu(self)
        menu.addAction("All")
        if vals:
            menu.addSeparator()
            MAX_ITEMS = 500
            for v in vals[:MAX_ITEMS]:
                menu.addAction(v)

        header = self.table.horizontalHeader()
        x = header.sectionViewportPosition(logicalIndex)
        y = header.height()
        global_pos = header.viewport().mapToGlobal(QPoint(x, y))
        action = menu.exec(global_pos)

        if action:
            text = action.text()
            if text == "All":
                if hasattr(self, "order_value_cb") and self.order_value_cb.isVisible():
                    self.order_value_cb.blockSignals(True)
                    self.order_value_cb.setCurrentIndex(0)
                    self.order_value_cb.blockSignals(False)
                self.order_text_search.setText("")
            else:
                if hasattr(self, "order_value_cb") and self.order_value_cb.isVisible():
                    idx = self.order_value_cb.findText(text)
                    if idx == -1:
                        self.order_value_cb.blockSignals(True)
                        self.order_value_cb.insertItem(1, text)
                        self.order_value_cb.setCurrentText(text)
                        self.order_value_cb.blockSignals(False)
                    else:
                        self.order_value_cb.setCurrentIndex(idx)
                self.order_text_search.setText(text)
            self._order_timer.start()

    def _on_order_col_changed(self, idx):
        data = self.order_col_cb.itemData(idx)
        if not data:
            typ = "text"
        else:
            _, typ = data
        is_number = (typ == "number")
        self.order_num_op.setVisible(is_number)
        self.order_num_input.setVisible(is_number)
        self.order_text_search.setVisible(not is_number)

        if typ == "text":
            col_key, _ = self.order_col_cb.itemData(idx)
            vals = self._unique_cache.get(col_key)
            if vals is None:
                vals = sorted({(str(getattr(o, col_key, "") or "")).strip() for o in self._orders if getattr(o, col_key, None) not in (None, "")})
                self._unique_cache[col_key] = vals
            self.order_value_cb.blockSignals(True)
            self.order_value_cb.clear()
            self.order_value_cb.addItem("All")
            for v in vals:
                self.order_value_cb.addItem(v)
            self.order_value_cb.blockSignals(False)
            self.order_value_cb.setVisible(True)
        else:
            self.order_value_cb.setVisible(False)

        self._order_timer.start()

    def _start_order_worker(self):
        idx = self.order_col_cb.currentIndex()
        col_key, col_type = self.order_col_cb.itemData(idx)
        val = None
        exact = False
        if col_type == "number":
            val = self._validate_and_parse_number(self.order_num_input)

        if col_type == "text" and self.order_value_cb.isVisible():
            sel = self.order_value_cb.currentText()
            if sel and sel != "All":
                exact = True
                keyword = sel.strip().lower()
            else:
                keyword = self.order_text_search.text().strip().lower()
        else:
            keyword = self.order_text_search.text().strip().lower()

        params = {
            "col_key": col_key,
            "col_type": col_type,
            "op": self.order_num_op.currentText(),
            "value": val,
            "keyword": keyword,
            "category": None,
            "exact": exact,
        }
        worker = FilterWorker(self._orders, "orders", params)
        worker.finished.connect(lambda res, w=worker: self._on_order_worker_finished_with_cleanup(res, w))
        worker.error.connect(self._on_worker_error)
        self._running_workers.append(worker)
        worker.start()

    def _on_order_worker_finished_with_cleanup(self, filtered, worker):
        try:
            self._on_order_worker_finished(filtered)
        finally:
            try:
                worker.quit()
            except Exception:
                pass
            worker.deleteLater()
            self._running_workers = [w for w in self._running_workers if w.isRunning()]

    def _on_order_worker_finished(self, filtered):
        self._fill_order_table(filtered)
        self.lbl_status.setText(f"{len(filtered)} pesanan ditampilkan.")

    def _build_tab_products(self):
        tab = QWidget(); tab.setObjectName("tabProduk")
        root = QVBoxLayout(tab); root.setContentsMargins(16,14,16,14); root.setSpacing(8)

        top = QHBoxLayout(); top.setSpacing(8)

        self.prod_col_cb = QComboBox()
        PROD_FILTER_COLS = [
            ("Product Name", "product_name", "text"),
            ("Category", "category", "category"),
            ("Price", "price", "number"),
        ]
        for label, key, typ in PROD_FILTER_COLS:
            self.prod_col_cb.addItem(label, (key, typ))
        self.prod_col_cb.currentIndexChanged.connect(self._on_prod_col_changed)
        top.addWidget(QLabel("Kolom:")); top.addWidget(self.prod_col_cb)

        self.prod_cat_cb = QComboBox(); self.prod_cat_cb.addItem("All"); self.prod_cat_cb.currentTextChanged.connect(lambda _: self._prod_timer.start())
        top.addWidget(self.prod_cat_cb)

        self.prod_num_op = QComboBox(); self.prod_num_op.addItems(["All","<","=" ,">"])
        self.prod_num_op.setMinimumWidth(80); self.prod_num_op.currentTextChanged.connect(lambda _: self._prod_timer.start())
        self.prod_num_input = QLineEdit(); self.prod_num_input.setPlaceholderText("Angka")
        self.prod_num_input.setObjectName("inputField")

        self.prod_num_input.textChanged.connect(lambda _: (self._validate_and_parse_number(self.prod_num_input), self._prod_timer.start()))
        top.addWidget(self.prod_num_op); top.addWidget(self.prod_num_input)

        top.addItem(QSpacerItem(20, 10, QSP.Expanding, QSP.Minimum))

        self.prod_text_search = QLineEdit(); self.prod_text_search.setPlaceholderText("Kata kunci (teks)")
        self.prod_text_search.setObjectName("inputField")
        self.prod_text_search.textChanged.connect(lambda _: self._prod_timer.start())
        self._prod_timer.timeout.connect(self._start_prod_worker)
        top.addWidget(self.prod_text_search)

        btn_help_p = QToolButton(); btn_help_p.setText("?"); btn_help_p.setToolTip("Petunjuk penggunaan filter dan kata kunci")
        btn_help_p.clicked.connect(lambda: self._show_help("products"))
        top.addWidget(btn_help_p)

        btn_add = QPushButton("Tambah Produk"); btn_add.setObjectName("btnPrimary"); btn_add.clicked.connect(self._tambah_produk)
        top.addWidget(btn_add)

        self.btn_edit_prod = QPushButton("Edit"); self.btn_edit_prod.setEnabled(False); self.btn_edit_prod.clicked.connect(self._edit_produk)
        self.btn_edit_prod.setObjectName("btnSecondary"); top.addWidget(self.btn_edit_prod)
        self.btn_del_prod = QPushButton("Hapus"); self.btn_del_prod.setEnabled(False); self.btn_del_prod.clicked.connect(self._hapus_produk)
        self.btn_del_prod.setObjectName("btnSecondary"); top.addWidget(self.btn_del_prod)

        root.addLayout(top)

        self.tbl_prod = QTableWidget(); self.tbl_prod.setObjectName("dataTable")
        pcols = ["ID", "Product Name", "Category", "Price"]
        self.tbl_prod.setColumnCount(len(pcols))
        self.tbl_prod.setHorizontalHeaderLabels(pcols)
        hdr = self.tbl_prod.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Stretch)
        hdr.setSectionResizeMode(0, QHeaderView.Fixed)
        self.tbl_prod.setColumnWidth(0, 50)
   
        hdr.sectionClicked.connect(self._on_prod_header_clicked)
        self.tbl_prod.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_prod.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_prod.setAlternatingRowColors(True)
        self.tbl_prod.verticalHeader().setVisible(False)
        self.tbl_prod.setShowGrid(False)
        self.tbl_prod.setStyleSheet("QTableWidget::item:selected { background-color: #3b82f6; color: white; }")
        self.tbl_prod.itemSelectionChanged.connect(self._on_prod_sel_changed)
        root.addWidget(self.tbl_prod)
        self.tabs.addTab(tab, "Produk")

        self._on_prod_col_changed(0)

    def _on_prod_header_clicked(self, logicalIndex: int):
        header_to_combo = {1: 0, 2: 1, 3: 2}
        if logicalIndex not in header_to_combo:
            return

        combo_idx = header_to_combo[logicalIndex]
        try:
            self.prod_col_cb.blockSignals(True)
            self.prod_col_cb.setCurrentIndex(combo_idx)
        finally:
            self.prod_col_cb.blockSignals(False)

        self._on_prod_col_changed(combo_idx)

        if logicalIndex != 2:
            return

        col_key = "category"
        vals = self._unique_cache.get(col_key)
        if vals is None:
            if hasattr(self, "_products") and getattr(self, "_products", None):
                vals = sorted({(str(p.category or "")).strip() for p in self._products if p.category})
            else:
                vals = sorted({self.tbl_prod.item(r, 2).text().strip() for r in range(self.tbl_prod.rowCount()) if self.tbl_prod.item(r, 2) and self.tbl_prod.item(r, 2).text().strip()})
            self._unique_cache[col_key] = vals

        menu = QMenu(self)
        menu.addAction("All")
        if vals:
            menu.addSeparator()
            for v in vals:
                menu.addAction(v)

        header = self.tbl_prod.horizontalHeader()
        x = header.sectionViewportPosition(logicalIndex)
        y = header.height()
        global_pos = header.viewport().mapToGlobal(QPoint(x, y))
        action = menu.exec(global_pos)

        if action:
            text = action.text()
            if text == "All":
                self.prod_cat_cb.blockSignals(True)
                self.prod_cat_cb.setCurrentIndex(0)
                self.prod_cat_cb.blockSignals(False)
            else:
                idx = self.prod_cat_cb.findText(text)
                if idx == -1:
                    self.prod_cat_cb.blockSignals(True)
                    self.prod_cat_cb.addItem(text)
                    self.prod_cat_cb.setCurrentText(text)
                    self.prod_cat_cb.blockSignals(False)
                else:
                    self.prod_cat_cb.setCurrentIndex(idx)
            self._prod_timer.start()


    def _on_prod_col_changed(self, idx):
        _, typ = self.prod_col_cb.itemData(idx)
        self.prod_cat_cb.setVisible(typ == "category")
        self.prod_num_op.setVisible(typ == "number")
        self.prod_num_input.setVisible(typ == "number")
        self.prod_text_search.setVisible(typ == "text")
        if typ == "category":
            prods = getattr(self.api, "_products_cache", []) or []
            cats = sorted({(p.category or "").strip() for p in prods if (p.category or "").strip()})
            self.prod_cat_cb.blockSignals(True)
            self.prod_cat_cb.clear()
            self.prod_cat_cb.addItem("All")
            for c in cats: self.prod_cat_cb.addItem(c)
            self.prod_cat_cb.blockSignals(False)
        self._prod_timer.start()

    def _start_prod_worker(self):
        prods = getattr(self.api, "_products_cache", []) or []
        idx = self.prod_col_cb.currentIndex()
        col_key, col_type = self.prod_col_cb.itemData(idx)
        val = None
        if col_type == "number":
            val = self._validate_and_parse_number(self.prod_num_input)
        params = {
            "col_key": col_key,
            "col_type": col_type,
            "op": self.prod_num_op.currentText(),
            "value": val,
            "keyword": self.prod_text_search.text().strip().lower(),
            "category": self.prod_cat_cb.currentText() if hasattr(self, "prod_cat_cb") else "All"
        }
        worker = FilterWorker(prods, "products", params)
        worker.finished.connect(lambda res, w=worker: self._on_prod_worker_finished_with_cleanup(res, w))
        worker.error.connect(self._on_worker_error)
        self._running_workers.append(worker)
        worker.start()

    def _on_prod_worker_finished_with_cleanup(self, filtered, worker):
        try:
            self._on_prod_worker_finished(filtered)
        finally:
            try:
                worker.quit()
            except Exception:
                pass
            worker.deleteLater()
            self._running_workers = [w for w in self._running_workers if w.isRunning()]

    def _on_prod_worker_finished(self, filtered):
        self.tbl_prod.setRowCount(0)
        for p in filtered:
            r = self.tbl_prod.rowCount(); self.tbl_prod.insertRow(r)
            for c, v in enumerate([str(p.id), p.product_name, p.category, Formatter.currency(p.price)]):
                item = QTableWidgetItem(v); item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont("Segoe UI", 10))
                self.tbl_prod.setItem(r, c, item)
        self.lbl_status.setText(f"{len(filtered)} produk ditampilkan.")

    def _build_tab_prediksi(self):
        self._tab_prediksi = TabPrediksi()
        self.tabs.addTab(self._tab_prediksi, "Prediksi ML")

    def _build_statusbar(self):
        sb = self.statusBar(); sb.setObjectName("statusBar")
        sb.setSizeGripEnabled(False)
        sb.setStyleSheet("QStatusBar::item { border: none; }")
        self.lbl_status = QLabel("  Memuat data...")
        sb.addWidget(self.lbl_status)
        lbl_anggota = QLabel(f"  👥  {_ANGGOTA_STR}  ")
        lbl_anggota.setStyleSheet("color:#6b7280;font-size:10px;")
        lbl_anggota.setToolTip("Nama & NIM anggota kelompok (tidak dapat diedit)")
        sb.addPermanentWidget(lbl_anggota)

    def refresh_all(self):
        self.lbl_status.setText("Mengambil data dari API...")
        
        self._loader = DataLoader(self.api)
        self._loader.finished.connect(self._on_data_loaded_with_cleanup)
        self._loader.error.connect(self._on_data_error)
        self._loader.start()
        self._load_products()

    def _on_data_loaded_with_cleanup(self, orders):
        try:
            self._on_data_loaded(orders)
        finally:
            try:
                self._loader.quit()
            except Exception:
                pass
            self._loader.deleteLater()
            self._loader = None

    @Slot(list)
    def _on_data_loaded(self, orders):
        self._orders = orders
        self._unique_cache.clear()
        self._start_order_worker()
        stats = self.api.summary_stats(orders)
        self._set_card(self._card_tx,    Formatter.number(stats["total_tx"]))
        self._set_card(self._card_rev,   Formatter.short_currency(stats["total_rev"]))
        self._set_card(self._card_avg,   Formatter.short_currency(stats["avg_rev"]))
        self._set_card(self._card_units, Formatter.number(stats["total_units"]))
        self._dash_chart.set_data({
            "revenue_by_kategori": self.api.revenue_by_kategori(orders),
            "revenue_by_status":   self.api.revenue_by_status(orders),
            "revenue_by_month":    self.api.revenue_by_month(orders),
            "units_by_payment":    self.api.units_by_payment(orders),
            "top_products":        self.api.top_products(orders),
        })
        try:
            self._tab_prediksi.load_orders(orders)
        except Exception:
            pass
        self.lbl_status.setText(f"{len(orders)} pesanan  |  diperbarui dari API")

    @Slot(str)
    def _on_data_error(self, msg):
        self.lbl_status.setText(f"Gagal memuat: {msg}")
        QMessageBox.warning(self, "Koneksi API Gagal", f"Detail: {msg}")

    def _fill_order_table(self, orders):
        STATUS_BG = {
            "Selesai":"#dcfce7","Diproses":"#dbeafe",
            "Dikirim":"#e0e7ff","Pending":"#fef9c3","Dibatalkan":"#fee2e2",
        }
        self.table.setRowCount(0)
        for o in orders:
            r = self.table.rowCount(); self.table.insertRow(r)
            vals = [
                str(o.id),
                o.order_id,
                o.customer_name,
                o.product_name,
                o.category,
                Formatter.currency(o.price),
                str(o.quantity),
                Formatter.currency(o.discount),
                Formatter.currency(o.total),
                Formatter.currency(o.shipping_fee),
                Formatter.currency(o.total_sales),
                o.status,
                o.shipping_address,
                o.customer_gender,
                o.customer_city,
                o.payment_method,
                o.courier,
                str(o.estimated_delivery_days),
                o.sales_channel,
                str(o.customer_rating),
                o.sales_date,
            ]
            bg = STATUS_BG.get(o.status)
            for c, v in enumerate(vals):
                item = QTableWidgetItem(v)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont("Segoe UI", 10))
                if bg: item.setBackground(QColor(bg))
                self.table.setItem(r, c, item)

    def _load_products(self):
        prods = self.api.get_products()
        cats = sorted({(p.category or "").strip() for p in prods if (p.category or "").strip()})
        self.prod_cat_cb.blockSignals(True)
        self.prod_cat_cb.clear()
        self.prod_cat_cb.addItem("All")
        for c in cats: self.prod_cat_cb.addItem(c)
        self.prod_cat_cb.blockSignals(False)
        self._start_prod_worker()

    def _on_prod_sel_changed(self):
        has = bool(self.tbl_prod.selectedItems())
        try:
            self.btn_edit_prod.setEnabled(has)
            self.btn_del_prod.setEnabled(has)
        except Exception:
            pass

    def _on_sel_changed(self):
        try:
            has = bool(self.table.selectedItems())
            self.btn_edit.setEnabled(has); self.btn_del.setEnabled(has)
        except Exception:
            pass

    def _selected_product(self) -> ProductRecord | None:
        row = self.tbl_prod.currentRow()
        if row < 0: return None
        pid = int(self.tbl_prod.item(row, 0).text())
        for p in getattr(self.api, "_products_cache", []) or []:
            if p.id == pid: return p
        return None

    def _selected_order(self):
        row = self.table.currentRow()
        if row < 0: return None
        oid = int(self.table.item(row, 0).text())
        for o in self._orders:
            if o.id == oid: return o
        return None

    def _tambah_produk(self):
        dlg = DialogProduk(self)
        if dlg.exec() != QDialog.Accepted: return
        try:
            self.api.create_product(dlg.get_data())
            self._load_products()
            self.lbl_status.setText("Produk baru berhasil ditambahkan.")
        except Exception as e:
            QMessageBox.critical(self, "Gagal Tambah Produk", str(e))

    def _edit_produk(self):
        p = self._selected_product()
        if not p: return
        dlg = DialogProduk(self, product=p)
        if dlg.exec() != QDialog.Accepted: return
        try:
            self.api.update_product(p.id, dlg.get_data())
            self._load_products()
            self.lbl_status.setText(f"Produk ID {p.id} berhasil diperbarui.")
        except Exception as e:
            QMessageBox.critical(self, "Gagal Edit Produk", str(e))

    def _hapus_produk(self):
        p = self._selected_product()
        if not p: return
        if QMessageBox.question(
            self, "Hapus Produk",
            f"Hapus produk '{p.product_name}'?\nPesanan yang menggunakan produk ini mungkin terpengaruh.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        ) != QMessageBox.Yes: return
        try:
            self.api.delete_product(p.id)
            self._load_products()
            self.lbl_status.setText(f"Produk '{p.product_name}' dihapus.")
        except Exception as e:
            QMessageBox.critical(self, "Gagal Hapus Produk", str(e))

    def _tambah_order(self):
        prods = self.api.get_products()
        if not prods:
            QMessageBox.warning(self, "Produk Kosong",
                                "Tambahkan produk dulu di tab Produk."); return
        dlg = DialogOrder(self, products=prods)
        if dlg.exec() != QDialog.Accepted: return
        try:
            self.api.create_order(dlg.get_raw_data())
            self.refresh_all()
            self.lbl_status.setText("Pesanan baru berhasil dicatat.")
        except Exception as e:
            QMessageBox.critical(self, "Gagal Tambah", str(e))

    def _edit_order(self):
        o = self._selected_order()
        if not o:
            QMessageBox.information(self, "Info", "Pilih baris yang ingin diedit."); return
        prods = self.api.get_products()
        dlg = DialogOrder(self, record=o, products=prods)
        if dlg.exec() != QDialog.Accepted: return
        try:
            self.api.update_order(o.id, dlg.get_raw_data())
            self.refresh_all()
            self.lbl_status.setText(f"Pesanan ID {o.id} berhasil diperbarui.")
        except Exception as e:
            QMessageBox.critical(self, "Gagal Update", str(e))

    def _hapus_order(self):
        o = self._selected_order()
        if not o:
            QMessageBox.information(self, "Info", "Pilih baris yang ingin dihapus."); return
        if QMessageBox.question(
            self, "Konfirmasi Hapus",
            f"Hapus pesanan ID {o.id} — {o.customer_name} ({o.product_name})?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        ) != QMessageBox.Yes: return
        try:
            self.api.delete_order(o.id)
            self.refresh_all()
            self.lbl_status.setText(f"Pesanan ID {o.id} dihapus.")
        except Exception as e:
            QMessageBox.critical(self, "Gagal Hapus", str(e))

    def _import_file(self):
        dlg = DialogImport(self, api=self.api)
        if dlg.exec() == QDialog.Accepted:
            self.refresh_all()
            self.lbl_status.setText("Import selesai.")

    # ─── Export CSV dan EXCEL

    def _get_visible_orders(self) -> list:
        visible_ids = set()
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item:
                try:
                    visible_ids.add(int(item.text()))
                except ValueError:
                    pass
        return [o for o in self._orders if o.id in visible_ids]

    def _export_headers(self):
        return [
            "ID DB", "Order ID", "Customer Name", "Product Name", "Category",
            "Price", "Quantity", "Discount", "Total", "Shipping Fee", "Total Sales",
            "Status", "Shipping Address", "Gender", "City", "Payment Method",
            "Courier", "Est. Delivery Days", "Sales Channel", "Rating", "Sales Date",
        ]

    def _order_to_row(self, o):
        return [
            o.id, o.order_id, o.customer_name, o.product_name, o.category,
            o.price, o.quantity, o.discount, o.total, o.shipping_fee, o.total_sales,
            o.status, o.shipping_address, o.customer_gender, o.customer_city,
            o.payment_method, o.courier, o.estimated_delivery_days,
            o.sales_channel, o.customer_rating, o.sales_date,
        ]

    def _export_csv(self):
        orders = self._get_visible_orders()
        if not orders:
            QMessageBox.warning(self, "Tidak Ada Data", "Tidak ada data yang bisa diekspor."); return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path, _ = QFileDialog.getSaveFileName(
            self, "Simpan sebagai CSV",
            f"GriyaData_Export_{ts}.csv",
            "CSV Files (*.csv)")
        if not path: return

        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(self._export_headers())
                for o in orders:
                    writer.writerow(self._order_to_row(o))
            self.lbl_status.setText(f"  CSV berhasil disimpan — {len(orders)} baris → {path}")
            QMessageBox.information(self, "Export Berhasil",
                                    f"{len(orders)} baris diekspor ke:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Gagal", str(e))

    def _export_excel(self):
        orders = self._get_visible_orders()
        if not orders:
            QMessageBox.warning(self, "Tidak Ada Data", "Tidak ada data yang bisa diekspor."); return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path, _ = QFileDialog.getSaveFileName(
            self, "Simpan sebagai Excel",
            f"GriyaData_Export_{ts}.xlsx",
            "Excel Files (*.xlsx)")
        if not path: return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Pesanan"

            # Header styling
            hdr_fill = PatternFill("solid", fgColor="1E3A5F")
            hdr_font = XLFont(bold=True, color="FFFFFF", size=11)
            headers  = self._export_headers()
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22

            # Data rows — zebra striping
            fill_even = PatternFill("solid", fgColor="F0F9FF")
            for row_i, o in enumerate(orders, 2):
                for col, val in enumerate(self._order_to_row(o), 1):
                    cell = ws.cell(row=row_i, column=col, value=val)
                    cell.alignment = Alignment(horizontal="center")
                    if row_i % 2 == 0:
                        cell.fill = fill_even

            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            wb.save(path)
            self.lbl_status.setText(f"  Excel berhasil disimpan — {len(orders)} baris → {path}")
            QMessageBox.information(self, "Export Berhasil",
                                    f"{len(orders)} baris diekspor ke:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Gagal", str(e))

    def _show_about(self):
        anggota_lines = "\n".join(f"  • {n}  —  NIM {nim}" for n, nim in ANGGOTA)
        QMessageBox.information(
            self, "Tentang GriyaData",
            f"GriyaData — Aplikasi Manajemen Penjualan Furniture\n\n"
            f"Versi  : 2.0\n"
            f"Backend: Railway (FastAPI + Supabase PostgreSQL)\n"
            f"ML     : Random Forest Regressor (scikit-learn)\n\n"
            f"Anggota Kelompok:\n{anggota_lines}"
        )

    def _logout(self):
        if QMessageBox.question(self, "Logout", "Yakin ingin keluar?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No) == QMessageBox.Yes:
            self.close()
            try:
                from ui.login_window import LoginWindow
                self._lw = LoginWindow(); self._lw.show()
            except Exception:
                pass

    def _on_worker_error(self, msg):
        self.lbl_status.setText(f"Filter error: {msg}")

    def _show_help(self, tab: str):
        if tab == "orders":
            text = (
                "Petunjuk Pencarian Pesanan\n\n"
                "- Pilih kolom yang ingin dicari dari dropdown 'Kolom'.\n"
                "- Jika kolom bertipe teks (mis. Pelanggan, Produk, Kota), ketik kata kunci di kotak 'Kata kunci' atau pilih nilai dari dropdown di sebelah kolom untuk exact-match.\n"
                "- Untuk kolom enum (status, category, payment_method, courier, sales_channel, customer_city, product_name), klik header kolom untuk menampilkan daftar nilai yang tersedia.\n"
                "- Pencarian teks bersifat case-insensitive dan mencari substring kecuali Anda memilih nilai dari dropdown (exact-match).\n"
                "- Jika kolom bertipe numerik (mis. Price, Total Sales, Qty), pilih operator (<, =, >) lalu masukkan angka.\n"
                "- Jika input angka berwarna merah, berarti tidak valid dan filter numerik akan diabaikan.\n"
                "- Hasil akan muncul otomatis setelah 300 ms dari input terakhir.\n"
                "- Gunakan tombol Refresh untuk memuat ulang data dari API."
            )
        else:
            text = (
                "Petunjuk Pencarian Produk\n\n"
                "- Pilih kolom yang ingin difilter: Product Name, Category, atau Price.\n"
                "- Untuk Category: pilih kategori dari dropdown atau klik header kolom 'Category' untuk daftar nilai.\n"
                "- Untuk Price: pilih operator (<, =, >) lalu masukkan angka.\n"
                "- Untuk Product Name: ketik kata kunci (substring, case-insensitive).\n"
                "- Jika input angka berwarna merah, berarti tidak valid dan filter numerik akan diabaikan.\n"
                "- Hasil akan muncul otomatis setelah 300 ms dari input terakhir.\n"
                "- Tombol Refresh memuat ulang daftar produk dari API."
            )
        QMessageBox.information(self, "Petunjuk Filter & Pencarian", text)